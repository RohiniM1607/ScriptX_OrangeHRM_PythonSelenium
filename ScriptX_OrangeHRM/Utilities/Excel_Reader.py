import openpyxl

def get_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_columns + 1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)
    return final_list

def invalid_AddLeave_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = 2
    total_columns = sheet.max_column

    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(2, total_columns + 1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)

    return final_list

def exceed_LeaveLimit(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = 3
    total_columns = sheet.max_column

    for r in range(3, total_rows + 1):
        row_list = []
        for c in range(1, total_columns + 1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)

    return final_list
def get_filtered_data(path, sheet_name, filter_value):
    data = get_data(path, sheet_name)
    return [row[1:]   for row in data if row[0] == filter_value]

def get_contact_details_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    for r in range(2, sheet.max_row + 1): 
        row_list = []
        for c in range(1, 11): 
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)
    return final_list

